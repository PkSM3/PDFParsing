

from openpyxl import load_workbook

class ExtractCSV:

	def __init__(self,filename):
		self.fn = filename
		self.heads = {}

	def startReading(self, prikey):


		wb = load_workbook(filename = self.fn , use_iterators = True)

		sheetname = wb.get_sheet_names()[0]

		ws = wb.get_sheet_by_name(name = sheetname)

		items = self.heads

		finalregs = []

		for row in ws.iter_rows():
			reg = {}
			for cell in row:
				if cell.row==1:
					items[cell.column] = cell.value
				else: 
					# print cell.column,cell.value
					if cell.column:
						reg[cell.column] = cell.value
			if reg.has_key(prikey) and reg[prikey]!=None and reg[prikey]!="":
				finalregs.append(reg)
			# else:
			#	print reg
				
		return finalregs


	def makeArray(self, results):

		finalarray = {}
		for i in results:

			ACR=""
			if i.has_key("A") and i["A"]!=None: ACR = i["A"].replace("\n","").replace("\t","").encode('utf-8')

			PI = ""
			if i.has_key("B") and i["B"]!=None: PI = i["B"].replace("\n","").replace("\t","").encode('utf-8')

			ID = i["C"].replace("\n","").replace("\t","").encode('utf-8')

			AG=""
			if i.has_key("D") and i["D"]!=None: AG = i["D"].replace("\n","").replace("\t","").encode('utf-8')

			TI=""
			if i.has_key("E") and i["E"]!=None: TI = i["E"].replace("\n","").replace("\t","").encode('utf-8')

			finalarray["BPM"+str(ID)] = {"PI":PI,"AG":AG,"ACR":ACR,"TI":TI}

		return finalarray


